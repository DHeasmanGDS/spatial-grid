"""Drill plan exporters — Excel workbook, shapefile bundle, surveys CSV."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from ..drill import DrillPlan

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="305496")
_LABEL_FONT = Font(bold=True)


# Shapefile DBF field names cap at 10 characters.
_COLLAR_RENAME = {
    "hole_name": "hole_name",
    "collar_e": "collar_e",
    "collar_n": "collar_n",
    "collar_rl": "collar_rl",
    "toe_e": "toe_e",
    "toe_n": "toe_n",
    "toe_rl": "toe_rl",
    "azimuth": "azimuth",
    "dip": "dip",
    "length_m": "length_m",
}
_SURVEY_RENAME = {
    "hole_name": "hole_name",
    "depth_m": "depth_m",
    "easting": "easting",
    "northing": "northing",
    "rl": "rl",
}
_TRACE_RENAME = {
    "hole_name": "hole_name",
    "length_m": "length_m",
    "azimuth": "azimuth",
    "dip": "dip",
}


def _write_header_row(ws, headers, row=1):
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def write_drill_excel(plan: DrillPlan, path: str | Path) -> Path:
    path = Path(path)
    wb = Workbook()
    spec = plan.spec

    # Parameters
    ws = wb.active
    ws.title = "Parameters"
    params = [
        ("Program name", spec.name),
        ("CRS", spec.crs),
        ("Survey interval (m)", spec.survey_interval_m),
        ("Cost per metre", spec.cost_per_metre if spec.cost_per_metre is not None else "—"),
    ]
    for r, (k, v) in enumerate(params, start=1):
        ws.cell(row=r, column=1, value=k).font = _LABEL_FONT
        ws.cell(row=r, column=2, value=v)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 26

    # Summary
    ws2 = wb.create_sheet("Summary")
    summary = [
        ("Hole count", plan.hole_count),
        ("Total metres", round(plan.total_metres, 2)),
        ("Total cost", round(plan.total_cost, 2) if plan.total_cost is not None else "—"),
        ("Mean hole length", round(plan.total_metres / plan.hole_count, 2)),
    ]
    for r, (k, v) in enumerate(summary, start=1):
        ws2.cell(row=r, column=1, value=k).font = _LABEL_FONT
        ws2.cell(row=r, column=2, value=v)
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 22

    # Holes
    ws3 = wb.create_sheet("Holes")
    cols = ["hole_name", "collar_e", "collar_n", "collar_rl",
            "toe_e", "toe_n", "toe_rl",
            "azimuth", "dip", "length_m"]
    _write_header_row(ws3, cols)
    for r, row in enumerate(plan.collars[cols].itertuples(index=False), start=2):
        for c, val in enumerate(row, start=1):
            ws3.cell(row=r, column=c, value=val)
    ws3.freeze_panes = "A2"
    for c in range(1, len(cols) + 1):
        ws3.column_dimensions[chr(64 + c)].width = 14

    # Surveys
    ws4 = wb.create_sheet("Surveys")
    survey_cols = ["hole_name", "depth_m", "easting", "northing", "rl"]
    _write_header_row(ws4, survey_cols)
    for r, row in enumerate(plan.surveys[survey_cols].itertuples(index=False), start=2):
        for c, val in enumerate(row, start=1):
            ws4.cell(row=r, column=c, value=val)
    ws4.freeze_panes = "A2"
    for c in range(1, len(survey_cols) + 1):
        ws4.column_dimensions[chr(64 + c)].width = 14

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def write_drill_shapefiles(
    plan: DrillPlan, output_dir: str | Path, basename: str
) -> tuple[Path, Path, Path]:
    """Write three shapefiles: collars (2D points), traces (3D linestrings),
    surveys (3D points). All in the program's CRS.
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    collars_path = output_dir / f"{basename}_collars.shp"
    traces_path = output_dir / f"{basename}_traces.shp"
    surveys_path = output_dir / f"{basename}_surveys.shp"

    plan.collars.rename(columns=_COLLAR_RENAME).to_file(collars_path, driver="ESRI Shapefile")
    plan.traces.rename(columns=_TRACE_RENAME).to_file(traces_path, driver="ESRI Shapefile")
    plan.surveys.rename(columns=_SURVEY_RENAME).to_file(surveys_path, driver="ESRI Shapefile")
    return collars_path, traces_path, surveys_path


def write_drill_csv(plan: DrillPlan, path: str | Path) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    plan.surveys.drop(columns=["geometry"]).to_csv(path, index=False, float_format="%.3f")
    return path
