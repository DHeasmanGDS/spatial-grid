"""Excel exporter — Parameters / Summary / Lines / Stations sheets."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from ..core import Grid

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="305496")
_LABEL_FONT = Font(bold=True)


def _write_header_row(ws, headers, row=1):
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def write_excel(grid: Grid, path: str | Path) -> Path:
    path = Path(path)
    wb = Workbook()

    # Parameters
    ws = wb.active
    ws.title = "Parameters"
    spec = grid.spec
    params = [
        ("Grid name", spec.grid_name),
        ("CRS", spec.crs),
        ("Anchor", spec.anchor),
        ("Reference easting", spec.centre_easting),
        ("Reference northing", spec.centre_northing),
        ("Azimuth (deg)", spec.azimuth_deg),
        ("Line spacing (m)", spec.line_spacing),
        ("Station spacing (m)", spec.station_spacing),
        ("Number of lines", spec.num_lines),
        ("Stations per line", spec.num_stations),
        ("Line naming", spec.line_naming),
        ("Station naming", spec.station_naming),
    ]
    for r, (k, v) in enumerate(params, start=1):
        ws.cell(row=r, column=1, value=k).font = _LABEL_FONT
        ws.cell(row=r, column=2, value=v)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 30

    # Summary
    ws2 = wb.create_sheet("Summary")
    summary = [
        ("Total stations", grid.total_stations),
        ("Total line-km", round(grid.total_line_km, 4)),
        ("Easting min", float(grid.stations["easting"].min())),
        ("Easting max", float(grid.stations["easting"].max())),
        ("Northing min", float(grid.stations["northing"].min())),
        ("Northing max", float(grid.stations["northing"].max())),
    ]
    for r, (k, v) in enumerate(summary, start=1):
        ws2.cell(row=r, column=1, value=k).font = _LABEL_FONT
        ws2.cell(row=r, column=2, value=v)
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 22

    # Stations
    ws3 = wb.create_sheet("Stations")
    cols = ["station_id", "line_id", "station_name", "line_offset_m",
            "station_offset_m", "easting", "northing"]
    _write_header_row(ws3, cols)
    df = grid.stations[cols]
    for r, row in enumerate(df.itertuples(index=False), start=2):
        for c, val in enumerate(row, start=1):
            ws3.cell(row=r, column=c, value=val)
    ws3.freeze_panes = "A2"
    for c in range(1, len(cols) + 1):
        ws3.column_dimensions[chr(64 + c)].width = 18

    # Lines
    ws4 = wb.create_sheet("Lines")
    line_cols = ["line_id", "line_offset_m", "length_m", "azimuth_deg"]
    _write_header_row(ws4, line_cols)
    for r, row in enumerate(grid.lines[line_cols].itertuples(index=False), start=2):
        for c, val in enumerate(row, start=1):
            ws4.cell(row=r, column=c, value=val)
    ws4.freeze_panes = "A2"
    for c in range(1, len(line_cols) + 1):
        ws4.column_dimensions[chr(64 + c)].width = 18

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
